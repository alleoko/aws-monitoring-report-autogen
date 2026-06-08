from openpyxl.styles import Font
from .common import (
    FILL_HEADER, FILL_SUBHDR, FILL_GREEN, FILL_YELLOW, FILL_ORANGE, FILL_RED,
    FONT_BOLD, FONT_BODY, BORDER_THIN, ALIGN_CENTER, ALIGN_LEFT,
    REGIONS, TS_LABEL, ACCOUNT_ID, ACCOUNT_NAME,
)


def build_summary_sheet(wb):
    ws = wb.create_sheet("Executive_Summary", 0)
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value     = "AWS INFRASTRUCTURE MONITORING REPORT"
    c.font      = Font(bold=True, color="FFFFFF", size=16)
    c.fill      = FILL_HEADER
    c.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 44

    ws.merge_cells("A2:J2")
    c = ws["A2"]
    c.value     = (f"Generated At: {TS_LABEL}   |   "
                   f"Account: {ACCOUNT_ID} ({ACCOUNT_NAME})   |   "
                   f"Regions: {', '.join(REGIONS)}")
    c.font      = Font(bold=True, color="FFFFFF", size=10)
    c.fill      = FILL_SUBHDR
    c.alignment = ALIGN_CENTER
    ws.row_dimensions[2].height = 20

    ws.cell(4, 1, "RISK COLOR LEGEND").font = Font(bold=True, size=10, color="1F4E79")
    ws.merge_cells("A4:C4")
    for r, (lbl, fill) in enumerate([
        ("Low  —  Healthy, no action needed",          FILL_GREEN),
        ("Medium  —  Minor warning, monitor closely",  FILL_YELLOW),
        ("High  —  Performance degradation / cost",    FILL_ORANGE),
        ("Critical  —  Immediate action required",     FILL_RED),
    ], start=5):
        ws.merge_cells(f"A{r}:C{r}")
        c = ws.cell(r, 1, f"   {lbl}")
        c.fill = fill
        c.font = Font(bold=True, size=9)
        c.alignment = ALIGN_LEFT
        c.border = BORDER_THIN
        ws.row_dimensions[r].height = 18

    ws.cell(4, 5, "REPORT CONTENTS").font = Font(bold=True, size=10, color="1F4E79")
    ws.merge_cells("E4:J4")
    sheets = [
        ("EC2_Health_Report",       "EC2 instances: CPU, memory, disk, security group exposure"),
        ("ECS_Cluster_Report",      "ECS clusters & services: task health, deployments, OOM"),
        ("RDS_Health_Report",       "RDS databases: storage, connections, backups, encryption"),
        ("Security_Posture_Report", "Security findings: open ports, IAM risks, S3 exposure, EBS"),
        ("Cost_And_Utilization",    "EC2 cost estimates, underutilized resources, savings potential"),
        ("CloudWatch_Alarms",       "All CloudWatch alarm states with severity and action"),
        ("LoadBalancer_Report",     "ALB/NLB: error rates, response times, unhealthy targets"),
        ("EBS_Volumes_Report",      "EBS encryption status, snapshot age, orphaned volumes"),
        ("AutoScaling_Report",      "ASG capacity, scaling policies, health check status"),
    ]
    for r, (name, desc) in enumerate(sheets, start=5):
        ws.cell(r, 5, name).font   = Font(bold=True, color="1F4E79", size=9)
        ws.cell(r, 5).border       = BORDER_THIN
        ws.cell(r, 6, "→").alignment = ALIGN_CENTER
        ws.merge_cells(f"G{r}:J{r}")
        ws.cell(r, 7, desc).font   = FONT_BODY
        ws.cell(r, 7).border       = BORDER_THIN
        ws.row_dimensions[r].height = 18

    for col, width in [("A", 30), ("B", 5), ("C", 10), ("D", 4),
                        ("E", 30), ("F", 4), ("G", 55)]:
        ws.column_dimensions[col].width = width
