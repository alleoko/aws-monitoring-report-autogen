from openpyxl.styles import Font
from .common import (
    REGIONS, NOW, START_24,
    get_session, cw_metric, tag, max_risk, risk_color,
    style_header_row, style_data_row, auto_width, title_banner, set_risk_cell,
)


def build_ec2_sheet(wb):
    ws = wb.create_sheet("EC2_Health_Report")
    headers = [
        "Instance ID", "Name", "Region", "Type", "State",
        "CPU Avg 24h %", "Memory % (Agent)", "Disk % (Agent)",
        "Net In (MB)", "Net Out (MB)",
        "Security Group(s)", "Open Ports Risk (0.0.0.0/0)",
        "EBS Optimized", "Launch Time",
        "Risk Level", "Recommendation",
    ]
    title_banner(ws, "EC2 Health Report", len(headers))
    style_header_row(ws, 2, len(headers))
    for i, h in enumerate(headers, 1):
        ws.cell(2, i, h)

    row    = 3
    rl_col = headers.index("Risk Level") + 1

    for region in REGIONS:
        sess = get_session(region)
        ec2  = sess.client("ec2")
        cw   = sess.client("cloudwatch")

        try:
            reservations = ec2.describe_instances(
                Filters=[{"Name": "instance-state-name",
                          "Values": ["running", "stopped", "impaired"]}]
            )["Reservations"]
        except Exception as e:
            style_data_row(ws, row, len(headers))
            ws.cell(row, 1, f"ERROR – {region}: {e}")
            row += 1
            continue

        sg_map = {}
        try:
            for sg in ec2.describe_security_groups()["SecurityGroups"]:
                sg_map[sg["GroupId"]] = sg
        except Exception:
            pass

        for res in reservations:
            for inst in res["Instances"]:
                iid   = inst["InstanceId"]
                iname = tag(inst.get("Tags"), "Name")
                itype = inst["InstanceType"]
                state = inst["State"]["Name"]

                dims = [{"Name": "InstanceId", "Value": iid}]
                cpu  = cw_metric(cw, "AWS/EC2",  "CPUUtilization",   dims)
                mem  = cw_metric(cw, "CWAgent",  "mem_used_percent", dims)
                disk = cw_metric(cw, "CWAgent",  "disk_used_percent",
                                 dims + [{"Name": "path", "Value": "/"},
                                         {"Name": "fstype", "Value": "xfs"}])
                if disk is None:
                    disk = cw_metric(cw, "CWAgent", "disk_used_percent",
                                     dims + [{"Name": "path", "Value": "/"},
                                             {"Name": "fstype", "Value": "ext4"}])

                net_in_raw  = cw_metric(cw, "AWS/EC2", "NetworkIn",  dims, stat="Sum")
                net_out_raw = cw_metric(cw, "AWS/EC2", "NetworkOut", dims, stat="Sum")
                net_in  = round(net_in_raw  / 1e6, 2) if net_in_raw  else None
                net_out = round(net_out_raw / 1e6, 2) if net_out_raw else None

                sg_names  = [sg.get("GroupName", sg["GroupId"]) for sg in inst.get("SecurityGroups", [])]
                open_risk = "No"
                for sg_ref in inst.get("SecurityGroups", []):
                    sg_data = sg_map.get(sg_ref["GroupId"], {})
                    for perm in sg_data.get("IpPermissions", []):
                        for ip in perm.get("IpRanges", []):
                            if ip.get("CidrIp") == "0.0.0.0/0":
                                fp = perm.get("FromPort", 0)
                                tp = perm.get("ToPort",   65535)
                                if fp not in (80, 443):
                                    open_risk = f"YES – port {fp}-{tp}"
                                    break

                ebs_opt    = "Yes" if inst.get("EbsOptimized") else "No"
                launch_str = inst.get("LaunchTime", NOW).strftime("%Y-%m-%d %H:%M")

                risk_level = "Low"
                rec_parts  = []

                if open_risk.startswith("YES"):
                    risk_level = "Critical"
                    rec_parts.append(f"SECURITY: {open_risk} open to 0.0.0.0/0. Restrict to VPN/bastion CIDR.")

                if cpu is not None and cpu >= 85:
                    risk_level = max_risk(risk_level, "Critical")
                    rec_parts.append(f"CPU critically high ({cpu}%) — risk of throttle/crash.")
                elif cpu is not None and cpu >= 70:
                    risk_level = max_risk(risk_level, "High")
                    rec_parts.append(f"CPU elevated ({cpu}%) — consider scale-up or load distribution.")
                elif cpu is not None and cpu >= 50:
                    risk_level = max_risk(risk_level, "Medium")
                    rec_parts.append(f"CPU moderate ({cpu}%) — monitor trend.")

                if mem is not None and mem > 85:
                    risk_level = max_risk(risk_level, "Critical")
                    rec_parts.append(f"Memory {mem}% — OOM risk imminent. Scale vertically or investigate leak.")
                elif mem is not None and mem > 70:
                    risk_level = max_risk(risk_level, "High")
                    rec_parts.append(f"Memory {mem}% — elevated. Investigate processes.")

                if disk is not None and disk > 85:
                    risk_level = max_risk(risk_level, "High")
                    rec_parts.append(f"Disk {disk}% — clean logs/tmp or expand volume.")
                elif disk is not None and disk > 70:
                    risk_level = max_risk(risk_level, "Medium")
                    rec_parts.append(f"Disk {disk}% — monitor for growth.")

                if state == "stopped":
                    risk_level = max_risk(risk_level, "Medium")
                    rec_parts.append("Instance stopped — EBS cost still accruing. Terminate if unused.")

                if ebs_opt == "No" and itype not in ("t2.micro", "t2.small", "t2.medium", "t1.micro"):
                    risk_level = max_risk(risk_level, "Medium")
                    rec_parts.append("EBS optimization disabled — enable for better I/O throughput.")

                if itype.startswith("t2."):
                    rec_parts.append("t2 generation is legacy — migrate to t3 for better performance/cost.")

                if not rec_parts:
                    rec_parts.append("Instance healthy — no action required.")

                data = [
                    iid, iname, region, itype, state.upper(),
                    cpu  if cpu  is not None else "N/A",
                    mem  if mem  is not None else "N/A (no agent)",
                    disk if disk is not None else "N/A (no agent)",
                    net_in  if net_in  is not None else "N/A",
                    net_out if net_out is not None else "N/A",
                    ", ".join(sg_names), open_risk,
                    ebs_opt, launch_str,
                    risk_level, " | ".join(rec_parts),
                ]
                style_data_row(ws, row, len(headers))
                for i, val in enumerate(data, 1):
                    ws.cell(row, i, val)
                set_risk_cell(ws, row, rl_col, risk_level)
                row += 1

    auto_width(ws)
    ws.freeze_panes = "A3"
    ws.row_dimensions[2].height = 30
