import boto3
import statistics
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime, timezone, timedelta

# ── Configuration ─────────────────────────────
# Set these to match your environment.
# Never hardcode AWS access keys here — use AWS CLI profiles instead.
PROFILE      = "your-aws-profile-name"
ACCOUNT_NAME = "Your Company Name"
REGIONS      = ["ap-southeast-1", "us-west-2"]
# ──────────────────────────────────────────────

NOW      = datetime.now(timezone.utc)
START_24 = NOW - timedelta(hours=24)
TS_LABEL = NOW.strftime("%Y-%m-%d %H:%M UTC")


def get_account_id():
    try:
        return boto3.Session(profile_name=PROFILE).client("sts").get_caller_identity()["Account"]
    except Exception:
        return "Unknown"


ACCOUNT_ID = get_account_id()

FILL_GREEN    = PatternFill("solid", fgColor="C6EFCE")
FILL_YELLOW   = PatternFill("solid", fgColor="FFEB9C")
FILL_ORANGE   = PatternFill("solid", fgColor="FFCC99")
FILL_RED      = PatternFill("solid", fgColor="FFC7CE")
FILL_HEADER   = PatternFill("solid", fgColor="1F4E79")
FILL_SUBHDR   = PatternFill("solid", fgColor="2E75B6")

FONT_HDR  = Font(bold=True, color="FFFFFF", size=10)
FONT_BOLD = Font(bold=True, size=9)
FONT_BODY = Font(size=9)

THIN  = Side(style="thin",   color="AAAAAA")
THICK = Side(style="medium", color="1F4E79")
BORDER_THIN  = Border(left=THIN,  right=THIN,  top=THIN,  bottom=THIN)
BORDER_THICK = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

RISK_FILL = {
    "Low":      FILL_GREEN,
    "Medium":   FILL_YELLOW,
    "High":     FILL_ORANGE,
    "Critical": FILL_RED,
    "LOW":      FILL_GREEN,
    "MEDIUM":   FILL_YELLOW,
    "HIGH":     FILL_ORANGE,
    "CRITICAL": FILL_RED,
    "OK":                FILL_GREEN,
    "ALARM":             FILL_RED,
    "INSUFFICIENT_DATA": FILL_YELLOW,
    "Pass":     FILL_GREEN,
    "Fail":     FILL_RED,
}

SEVERITY_FILL = {
    "Critical": FILL_RED,
    "High":     FILL_ORANGE,
    "Medium":   FILL_YELLOW,
    "Low":      FILL_GREEN,
}


def get_session(region):
    return boto3.Session(profile_name=PROFILE, region_name=region)


def cw_metric(cw, namespace, metric_name, dimensions, stat="Average", period=3600):
    try:
        resp   = cw.get_metric_statistics(
            Namespace=namespace,
            MetricName=metric_name,
            Dimensions=dimensions,
            StartTime=START_24,
            EndTime=NOW,
            Period=period,
            Statistics=[stat],
        )
        points = resp.get("Datapoints", [])
        if not points:
            return None
        return round(statistics.mean(p[stat] for p in points), 2)
    except Exception:
        return None


def safe(fn, default=None):
    try:
        return fn()
    except Exception:
        return default


def tag(tags, key):
    if not tags:
        return "—"
    for t in tags:
        if t.get("Key") == key:
            return t.get("Value", "—")
    return "—"


def max_risk(a, b):
    order = {"Low": 0, "Medium": 1, "High": 2, "Critical": 3}
    return a if order.get(a, 0) >= order.get(b, 0) else b


def risk_color(risk_str):
    return RISK_FILL.get(risk_str)


def style_header_row(ws, row_num, cols):
    from openpyxl.utils import get_column_letter
    for col in range(1, cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill      = FILL_HEADER
        cell.font      = FONT_HDR
        cell.alignment = ALIGN_CENTER
        cell.border    = BORDER_THICK


def style_data_row(ws, row_num, cols, fill=None):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font      = FONT_BODY
        cell.alignment = ALIGN_LEFT
        cell.border    = BORDER_THIN
        if fill:
            cell.fill = fill


def auto_width(ws, min_w=12, max_w=60):
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        length = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)


def title_banner(ws, title, col_span):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_span)
    c       = ws.cell(1, 1, f"  {title}   |   Report Generated At: {TS_LABEL}")
    c.fill  = FILL_SUBHDR
    c.font  = Font(bold=True, color="FFFFFF", size=11)
    c.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 22


def set_risk_cell(ws, row, col_idx, risk_level):
    cell      = ws.cell(row, col_idx)
    cell.fill = risk_color(risk_level) or FILL_GREEN
    cell.font = Font(bold=True, size=9)
