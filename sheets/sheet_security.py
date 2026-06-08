from openpyxl.styles import Font
from botocore.exceptions import ClientError
from .common import (
    REGIONS, NOW,
    get_session, safe, tag,
    SEVERITY_FILL, FILL_GREEN, FILL_RED,
    style_header_row, style_data_row, auto_width, title_banner,
)


def build_security_sheet(wb):
    ws = wb.create_sheet("Security_Posture_Report")
    headers = [
        "Resource Type", "Resource Name / ID", "Region",
        "Issue Type", "Severity", "Description",
        "Recommendation", "Compliance",
    ]
    title_banner(ws, "Security Posture Report", len(headers))
    style_header_row(ws, 2, len(headers))
    for i, h in enumerate(headers, 1):
        ws.cell(2, i, h)

    findings = []

    for region in REGIONS:
        sess = get_session(region)
        ec2  = sess.client("ec2")

        sgs = safe(lambda: ec2.describe_security_groups()["SecurityGroups"], [])
        for sg in sgs:
            sgid   = sg["GroupId"]
            sgname = sg.get("GroupName", sgid)
            for perm in sg.get("IpPermissions", []):
                fp    = perm.get("FromPort", 0)
                tp    = perm.get("ToPort",   65535)
                proto = perm.get("IpProtocol", "-1")
                for ip in perm.get("IpRanges", []):
                    if ip.get("CidrIp") == "0.0.0.0/0":
                        if fp in (22, 3389) or proto == "-1":
                            sev  = "Critical"
                            desc = f"SSH(22)/RDP(3389)/ALL exposed to 0.0.0.0/0 on {sgname}"
                            rec  = "Remove 0.0.0.0/0. Use SSM Session Manager or VPN for admin access."
                        elif fp not in (80, 443):
                            sev  = "High"
                            desc = f"Port {fp}-{tp} ({proto}) open to 0.0.0.0/0 on {sgname}"
                            rec  = f"Restrict port {fp} to known CIDR ranges."
                        else:
                            continue
                        findings.append(("Security Group", f"{sgname} ({sgid})", region,
                                         "Open Port / Public Exposure", sev, desc, rec, "Fail"))
                for ip6 in perm.get("Ipv6Ranges", []):
                    if ip6.get("CidrIpv6") == "::/0":
                        findings.append(("Security Group", f"{sgname} ({sgid})", region,
                                         "IPv6 Public Exposure", "High",
                                         f"IPv6 ::/0 on port {fp}-{tp}",
                                         "Restrict IPv6 CIDR to known ranges.", "Fail"))

        vols = safe(lambda: ec2.describe_volumes()["Volumes"], [])
        for v in vols:
            if not v.get("Encrypted"):
                vname = tag(v.get("Tags"), "Name")
                findings.append(("EBS Volume", f"{v['VolumeId']} ({vname})", region,
                                 "Encryption Disabled", "High",
                                 "EBS volume not encrypted at rest.",
                                 "Create encrypted snapshot, replace volume. Enable default EBS encryption in region.", "Fail"))

        if region == REGIONS[0]:
            s3 = get_session(region).client("s3")
            buckets = safe(lambda: s3.list_buckets().get("Buckets", []), [])
            for b in buckets:
                bname = b["Name"]
                try:
                    pab = s3.get_public_access_block(Bucket=bname)["PublicAccessBlockConfiguration"]
                    if not all(pab.values()):
                        findings.append(("S3 Bucket", bname, "global",
                                         "Public Access Enabled", "High",
                                         "S3 Block Public Access not fully enabled.",
                                         "Enable all 4 Block Public Access settings.", "Fail"))
                except ClientError as e:
                    if "NoSuchPublicAccessBlockConfiguration" in str(e):
                        findings.append(("S3 Bucket", bname, "global",
                                         "No Public Access Block", "Critical",
                                         "No Public Access Block config — bucket potentially public.",
                                         "Apply Block Public Access at bucket and account level.", "Fail"))

                try:
                    enc   = s3.get_bucket_encryption(Bucket=bname)
                    rules = enc.get("ServerSideEncryptionConfiguration", {}).get("Rules", [])
                    if not rules:
                        raise Exception()
                except Exception:
                    findings.append(("S3 Bucket", bname, "global",
                                     "Encryption Disabled", "Medium",
                                     "Default bucket encryption not configured.",
                                     "Enable SSE-S3 or SSE-KMS default encryption.", "Fail"))

            try:
                iam   = get_session(region).client("iam")
                users = iam.list_users()["Users"]
                for u in users:
                    uname = u["UserName"]
                    try:
                        iam.get_login_profile(UserName=uname)
                        mfa = iam.list_mfa_devices(UserName=uname)["MFADevices"]
                        if not mfa:
                            findings.append(("IAM User", uname, "global",
                                             "MFA Not Enabled", "Critical",
                                             "Console user without MFA — account takeover risk.",
                                             "Enforce MFA via IAM policy or AWS Organizations SCP.", "Fail"))
                    except ClientError:
                        pass

                    try:
                        keys = iam.list_access_keys(UserName=uname)["AccessKeyMetadata"]
                        for k in keys:
                            if k["Status"] == "Active":
                                age = (NOW - k["CreateDate"]).days
                                if age > 90:
                                    sev = "Critical" if age > 180 else "High"
                                    findings.append(("IAM User", uname, "global",
                                                     "Stale Access Key", sev,
                                                     f"Key {k['AccessKeyId'][:12]}... is {age} days old.",
                                                     "Rotate access keys every 90 days.", "Fail"))
                    except Exception:
                        pass
            except Exception:
                pass

    row = 3
    severity_order = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
    findings.sort(key=lambda x: severity_order.get(x[4], 9))

    if not findings:
        ws.cell(3, 1, "No security issues detected across scanned resources.")
    else:
        for f in findings:
            style_data_row(ws, row, len(headers))
            for i, val in enumerate(f, 1):
                ws.cell(row, i, val)
            ws.cell(row, 5).fill = SEVERITY_FILL.get(f[4], FILL_GREEN)
            ws.cell(row, 5).font = Font(bold=True, size=9)
            ws.cell(row, 8).fill = FILL_GREEN if f[7] == "Pass" else FILL_RED
            ws.cell(row, 8).font = Font(bold=True, size=9)
            row += 1

    auto_width(ws)
    ws.freeze_panes = "A3"
    ws.row_dimensions[2].height = 30
