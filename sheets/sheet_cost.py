from openpyxl.styles import Font
from .common import (
    REGIONS,
    get_session, cw_metric, tag,
    FILL_ORANGE, FONT_BOLD,
    style_header_row, style_data_row, auto_width, title_banner,
)

HOURLY_PRICE = {
    "t2.micro": 0.0116, "t2.small": 0.023, "t2.medium": 0.0464, "t2.large": 0.0928,
    "t3.micro": 0.0104, "t3.small": 0.0208, "t3.medium": 0.0416, "t3.large": 0.0832,
    "t3.xlarge": 0.1664, "t3.2xlarge": 0.3328,
    "m5.large": 0.096,  "m5.xlarge": 0.192,  "m5.2xlarge": 0.384,
    "c5.large": 0.085,  "c5.xlarge": 0.17,
    "r5.large": 0.126,  "r5.xlarge": 0.252,
}


def build_cost_sheet(wb):
    ws = wb.create_sheet("Cost_And_Utilization")
    headers = [
        "Service", "Resource ID / Name", "Region", "Type",
        "CPU Avg %", "State",
        "Underutilized", "Est. Monthly Cost (USD)",
        "Optimization Suggestion", "Savings Potential %",
    ]
    title_banner(ws, "Cost & Utilization Summary", len(headers))
    style_header_row(ws, 2, len(headers))
    for i, h in enumerate(headers, 1):
        ws.cell(2, i, h)

    row = 3
    total_monthly = 0.0
    total_savings  = 0.0

    for region in REGIONS:
        sess = get_session(region)
        ec2  = sess.client("ec2")
        cw   = sess.client("cloudwatch")

        try:
            reservations = ec2.describe_instances(
                Filters=[{"Name": "instance-state-name", "Values": ["running", "stopped"]}]
            )["Reservations"]
        except Exception:
            reservations = []

        for res in reservations:
            for inst in res["Instances"]:
                iid   = inst["InstanceId"]
                iname = tag(inst.get("Tags"), "Name")
                itype = inst["InstanceType"]
                state = inst["State"]["Name"]

                dims    = [{"Name": "InstanceId", "Value": iid}]
                cpu_avg = cw_metric(cw, "AWS/EC2", "CPUUtilization", dims)

                hourly  = HOURLY_PRICE.get(itype, 0.05)
                monthly = round(hourly * 24 * 30, 2)
                total_monthly += monthly

                underutilized = "No"
                suggestion    = "Instance appears healthy — maintain current configuration."
                savings_pct   = 0

                if state == "stopped":
                    underutilized = "YES"
                    suggestion    = "STOPPED — EBS still billed. Terminate if permanently unused."
                    savings_pct   = 100
                elif cpu_avg is not None and cpu_avg < 5:
                    underutilized = "YES"
                    suggestion    = f"CPU {cpu_avg}% avg — severely underutilized. Downsize type or use Savings Plan."
                    savings_pct   = 50
                elif cpu_avg is not None and cpu_avg < 15:
                    underutilized = "YES"
                    suggestion    = f"CPU {cpu_avg}% — low utilization. Consider smaller type or Reserved Instance."
                    savings_pct   = 35
                elif itype.startswith("t2."):
                    suggestion    = "t2 is legacy — migrate to t3 for ~30% better price-performance."
                    savings_pct   = 20

                savings_monthly = round(monthly * savings_pct / 100, 2)
                total_savings  += savings_monthly

                data = [
                    "EC2", f"{iname} ({iid})", region, itype,
                    cpu_avg if cpu_avg is not None else "N/A",
                    state.upper(), underutilized,
                    f"${monthly}",
                    suggestion, f"{savings_pct}%",
                ]
                style_data_row(ws, row, len(headers))
                for i, val in enumerate(data, 1):
                    ws.cell(row, i, val)
                if underutilized == "YES":
                    ws.cell(row, 7).fill = FILL_ORANGE
                    ws.cell(row, 7).font = Font(bold=True, size=9)
                row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.cell(row, 1, "TOTAL ESTIMATED MONTHLY COST (EC2 ONLY)").font = FONT_BOLD
    ws.cell(row, 8, f"${round(total_monthly, 2)}").font = Font(bold=True, color="1F4E79", size=11)
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.cell(row, 1, "POTENTIAL MONTHLY SAVINGS").font = FONT_BOLD
    ws.cell(row, 8, f"${round(total_savings, 2)}").font = Font(bold=True, color="375623", size=11)

    auto_width(ws)
    ws.freeze_panes = "A3"
    ws.row_dimensions[2].height = 30
