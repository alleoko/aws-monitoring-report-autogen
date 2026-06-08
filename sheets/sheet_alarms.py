from openpyxl.styles import Font
from .common import (
    REGIONS,
    get_session, risk_color,
    SEVERITY_FILL, FILL_YELLOW,
    style_header_row, style_data_row, auto_width, title_banner,
)


def build_alarms_sheet(wb):
    ws = wb.create_sheet("CloudWatch_Alarms")
    headers = [
        "Alarm Name", "Region", "Namespace", "Metric",
        "State", "Threshold / Condition",
        "Last State Update", "Actions Enabled",
        "Severity", "Action Required",
    ]
    title_banner(ws, "CloudWatch Alarms Status", len(headers))
    style_header_row(ws, 2, len(headers))
    for i, h in enumerate(headers, 1):
        ws.cell(2, i, h)

    row    = 3
    rl_col = headers.index("Severity") + 1

    for region in REGIONS:
        sess = get_session(region)
        cw   = sess.client("cloudwatch")
        try:
            paginator = cw.get_paginator("describe_alarms")
            for page in paginator.paginate():
                for alarm in page.get("MetricAlarms", []):
                    aname    = alarm["AlarmName"]
                    state    = alarm["StateValue"]
                    metric   = alarm.get("MetricName", "—")
                    ns       = alarm.get("Namespace", "—")
                    cond     = (f"{alarm.get('ComparisonOperator','?')} "
                                f"{alarm.get('Threshold','?')} for "
                                f"{alarm.get('EvaluationPeriods','?')} periods")
                    updated  = str(alarm.get("StateUpdatedTimestamp", "N/A"))[:19]
                    act_en   = "Yes" if alarm.get("ActionsEnabled") else "No"

                    if state == "ALARM":
                        severity = "Critical"
                        action   = "ACTIVE ALARM — investigate immediately. Verify SNS/actions are firing."
                    elif state == "INSUFFICIENT_DATA":
                        severity = "Medium"
                        action   = "No data — verify CloudWatch agent is publishing metrics."
                    else:
                        severity = "Low"
                        action   = "Alarm in OK state — no action required."

                    if act_en == "No" and state == "ALARM":
                        action += " ⚠ Alarm actions DISABLED — notifications not firing!"

                    data = [
                        aname, region, ns, metric,
                        state, cond, updated, act_en,
                        severity, action,
                    ]
                    style_data_row(ws, row, len(headers))
                    for i, val in enumerate(data, 1):
                        ws.cell(row, i, val)

                    state_cell = ws.cell(row, 5)
                    state_cell.fill = risk_color(state) or FILL_YELLOW
                    state_cell.font = Font(bold=True, size=9)

                    sev_cell = ws.cell(row, rl_col)
                    sev_cell.fill = SEVERITY_FILL.get(severity, FILL_YELLOW)
                    sev_cell.font = Font(bold=True, size=9)
                    row += 1
        except Exception as e:
            style_data_row(ws, row, len(headers))
            ws.cell(row, 1, f"Error in {region}: {e}")
            row += 1

    if row == 3:
        ws.cell(3, 1, "No CloudWatch alarms found.")

    auto_width(ws)
    ws.freeze_panes = "A3"
    ws.row_dimensions[2].height = 30
